from abc import abstractclassmethod
import openpyxl
import json
import sys


class ConvertToSomething():
    """
    Excel形式を変換する親クラス。

    Attributes
    ----------
    _sheet : Objct
        Excelのシート
    _output : FileObjct
        出力ファイル
    _columns : List of str
        カラム名のリスト
    """

    def __init__(self, sheet, output):
        """
        Parameters
        ----------
        _sheet : Objct
            Excelのシート
        _output : FileObjct
            出力ファイル
        """
        self._sheet = sheet
        self._output = output
        self._columns = []

    def _write_header(self):
        """
        ヘッダー出力処理
        """
        # ファイルがあれば上書きにするため、modeはwとする
        with open(self._output, mode='w') as f:
            f.write("[")

    def _write_footer(self):
        """
        フッター出力処理
        """
        with open(self._output, mode='a') as f:
            f.write("]")

    def _set_columns(self):
        """
        Excelのヘッダーを取得しcolumnsにセットする
        """
        column = 1
        while self._check(1, column):
            self._columns.append(self._sheet.cell(row=1, column=column).value)
            column += 1

    def _check(self, row, column):
        """
        指定のセルに値があるかチェックする

        Returns
        ----------
        bool
            値が入っていればTrue
        """
        data = self._sheet.cell(row=row, column=column).value
        if data != '' and data != None:
            return True
        return False

    def _write_body(self):
        """
        1レコードずつ読み取ってデータを出力する
        """
        row = 2
        while self._check(row, 1):
            # add comma
            if row != 2:
                with open(self._output, mode='a') as f:
                    f.write(",")

            # 対象行から各値を取得する
            data = {}
            for index, column in enumerate(self._columns):
                data[column] = self._sheet.cell(row=row, column=index+1).value

            self._write_data(data)

            row += 1

    @abstractclassmethod
    def _write_data(self, data):
        """
        各出力形式に則ってデータ本体を出力する
        """
        pass

    def convert(self):
        """
        出力処理を順番に実行していく
        """
        self._write_header()
        self._set_columns()
        self._write_body()
        self._write_footer()


class ConvertToJson(ConvertToSomething):
    """
    Excel形式をJsonへ変換するクラス

    Attributes
    ----------
    _sheet : Objct
        Excelのシート
    _output : FileObjct
        出力ファイル
    _columns : List of str
        カラム名のリスト
    """

    def _write_data(self, data):
        """
        Json形式で出力する
        """
        with open(self._output, mode='a') as f:
            f.write(json.dumps(data, sort_keys=False,
                               ensure_ascii=False, indent=4))


class ConvertToPhpArray(ConvertToSomething):
    """
    Excel形式をPHPの連想配へ変換するクラス

    Attributes
    ----------
    _sheet : Objct
        Excelのシート
    _output : FileObjct
        出力ファイル
    _columns : List of str
        カラム名のリスト
    """

    def _write_data(self, data):
        """
        PHPの連想配列として出力する
        """
        with open(self._output, mode='a') as f:
            f.write('[\n    ')
            for index, (key, value) in enumerate(data.items()):
                f.write(self.__add_quote(key)+' => ' +
                        self.__format_value(value))
                if index != len(data)-1:
                    f.write(',')
            f.write('\n]')

    def __add_quote(self, value):
        """
        値の左右に"をつける

        Returns
        ----------
        str
        """
        return '\"'+str(value)+'\"'

    def __format_value(self, value):
        """
        型に応じて出力形式を変える

        Returns
        ----------
        str
        """
        if isinstance(value, (int, float)):
            return str(value)
        else:
            return self.__add_quote(value)


def main():
    """
    メイン処理
    """
    # 設定ファイル読み込み
    with open('settings.json', 'r') as f:
        try:
            settings = json.load(f)

            book = openpyxl.load_workbook(settings['dataFile'])
            sheet = book[settings['dataSheet']]
            output = settings['outputFile']
            output_type = settings['outputFileType']
            if output_type not in ['json', 'php']:
                raise Exception('outputFileTypeにはjsonかphpを設定してください。')

        except Exception as e:
            print("settings.jsonを読み込めませんでした。形式を見直してください。")
            print(e)
            sys.exit(1)

    print(output_type+'の形式で出力します。')

    try:
        if output_type == 'json':
            converter = ConvertToJson(sheet, output)
        elif output_type == 'php':
            converter = ConvertToPhpArray(sheet, output)
    except Exception as e:
        print('converterを取得できませんでした。')
        print(e)
        sys.exit(1)

    try:
        converter.convert()
    except Exception as e:
        print('正常に出力できませんでした。')
        print(e)
        sys.exit(1)

    print('出力処理が完了しました。')
    sys.exit()


main()
